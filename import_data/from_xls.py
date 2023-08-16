import datetime
import logging
from datetime import date
from enum import Enum

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from abstract.load_data import LoadData, MetricsTuple
from db.model import Categories, DataSources, Types
from exceptions import InvalidXLSXDataStructure

# Оригинальные данные тестовые, без даты.
# Данная константа определяет, сколько строк XLS файла будут с одинаковой датой
ROWS_WITH_THE_SAME_DATE = 3


class RowTypes(Enum):
    """Rows templates of test xls file"""
    DATA = (int, str, int, int, int, int, int, int, int, int,)
    HEADER_CATEGORIES = ('id', 'company', 'fact', 'forecast',)
    HEADER_TYPES = ('Qliq', 'Qoil', 'Qliq', 'Qoil',)
    HEADER_DATASOURCES = ('data1', 'data2', 'data1', 'data2', 'data1', 'data2', 'data1', 'data2',)
    UNKNOWN = "Unknown"


class LoadFromXLSX(LoadData):

    source_file_name: str = None
    _workbook: Workbook = None
    _sheet: Worksheet = None
    _year: int = None
    _month: int = None
    _dqy: int = None
    _random_date = None

    # Cell number maps
    CELL_COMPANY = 1
    CELLS_FACT = (2, 3, 4, 5)
    CELLS_FORECAST = (6, 7, 8, 9)
    CELLS_QLIQ = (2, 3, 6, 7)
    CELLS_QOIL = (4, 5, 8, 9)
    CELLS_DATA1 = (2, 4, 6, 8)
    CELLS_DATA2 = (3, 5, 7, 9)

    def __init__(self, source_file_name: str, year: int, month: int):
        self.source_file_name = source_file_name
        self._year = year
        self._month = month
        self._random_date = datetime.date(year, month, 1)
        try:
            self._workbook = openpyxl.load_workbook(self.source_file_name)
        except (InvalidFileException, FileNotFoundError) as e:
            logging.error(e)
            raise e

        sheets: list = self._workbook.sheetnames
        self._sheet = self._workbook[sheets[0]]

    def get_data(self) -> list[MetricsTuple]:
        """Main function for loading data from xls"""
        rows = []
        first_day = datetime.date(self._year, self._month, 1)
        day_counter = 0
        rows_per_day_counter = 0
        for row in self._sheet.iter_rows(values_only=True):
            row_type = self._get_row_type(row)
            if row_type == RowTypes.DATA:
                row_date = first_day + datetime.timedelta(days=day_counter)
                rows.extend(self._parse_row(row, row_date))
                rows_per_day_counter += 1
                if rows_per_day_counter >= ROWS_WITH_THE_SAME_DATE:
                    rows_per_day_counter = 0
                    day_counter += 1
        if len(rows) == 0:
            raise InvalidXLSXDataStructure
        return rows

    @classmethod
    def _get_row_type(cls, row: tuple) -> RowTypes:
        """Get row type using cell number maps"""
        cleaned_row = tuple(item for item in row if item is not None)
        if all(isinstance(value, datatype) for value, datatype in zip(row, RowTypes.DATA.value)):
            return RowTypes.DATA
        elif cleaned_row == RowTypes.HEADER_CATEGORIES.value:
            return RowTypes.HEADER_CATEGORIES
        elif cleaned_row == RowTypes.HEADER_TYPES.value:
            return RowTypes.HEADER_TYPES
        elif cleaned_row == RowTypes.HEADER_DATASOURCES.value:
            return RowTypes.HEADER_DATASOURCES
        else:
            return RowTypes.UNKNOWN

    def _parse_row(self, row: tuple, row_date: date) -> list[MetricsTuple]:
        """Parses data row"""
        result: list[MetricsTuple] = []
        for num_cell in range(2, len(row)):
            metrics = MetricsTuple(
                date=row_date,
                company=row[self.CELL_COMPANY],
                category=self._get_category(num_cell),
                type=self._get_type(num_cell),
                data=self._get_data_source(num_cell),
                quantity=row[num_cell]
            )
            if metrics.quantity is not None:
                result.append(metrics)
        return result

    def _get_category(self, num_cell: int) -> Categories | None:
        if num_cell in self.CELLS_FACT:
            return Categories.fact
        elif num_cell in self.CELLS_FORECAST:
            return Categories.forecast
        else:
            return None

    def _get_type(self, num_cell: int) -> Types | None:
        if num_cell in self.CELLS_QLIQ:
            return Types.liq
        elif num_cell in self.CELLS_QOIL:
            return Types.oil
        else:
            return None

    def _get_data_source(self, num_cell: int) -> DataSources | None:
        if num_cell in self.CELLS_DATA1:
            return DataSources.data1
        elif num_cell in self.CELLS_DATA2:
            return DataSources.data2
        else:
            return None

    def __del__(self):
        if self._workbook is not None:
            self._workbook.close()
